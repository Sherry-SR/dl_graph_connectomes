import importlib

import os
import numpy as np
import pandas as pd
import itertools
from openpyxl import load_workbook
import pickle

import torch
from torch_geometric.utils import dense_to_sparse
from torch_geometric.data import Data, InMemoryDataset

def create_data_loader(config):
    # to be done
    pass

def read_xlsx(path):
    workbook = load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]
    data = sheet.values
    cols = next(data)[0:]
    data = list(data)
    data = (itertools.islice(r, 0, None) for r in data)
    df = pd.DataFrame(data, columns = cols)
    return df

def read_connectome(path, filelist, label = None, num_nodes = None):
    dataset = []
    for f in filelist:
        filepath = os.path.join(path, f)
        matrix = torch.tensor(np.loadtxt(filepath), dtype=torch.float)
        edge_index, value = dense_to_sparse(matrix)
        if label is not None:
            key = '_'.join(f.split('_')[0:2])
            y = label[key]
        else:
            y = None
        x = torch.ones([matrix.shape[0], 1], dtype=torch.float)
        if num_nodes is not None:
            if matrix.shape[0] != num_nodes:
                print("exclude:", f)
                continue
        dataset.append(Data(x = x, edge_index = edge_index, edge_attr = value, y = y))
    return dataset

def connectome_process(DATA_PATH, LABEL_PATH, OUTPUT_PATH, label_name, index_name = "Subject", factorize = False, output_name = None, num_nodes = None):
    filelist = os.listdir(DATA_PATH)
    labels = read_xlsx(LABEL_PATH)
    if output_name is None:
        output_name = os.path.basename(DATA_PATH)+"_"+label_name+".pkl"
    if factorize:
        labels[label_name], uniques = pd.factorize(labels[label_name])
        with open(os.path.join(OUTPUT_PATH, output_name.split('.')[0]+"_uniques.txt"), "w") as f:
            print("Label uniques:", uniques.values, file = f)
    label = dict([(i, v) for i, v in zip(labels[index_name], labels[label_name])])
    dataset = read_connectome(DATA_PATH, filelist, label, num_nodes)
    with open(os.path.join(OUTPUT_PATH, output_name), "wb") as f:
        pickle.dump(dataset, f)

class ConnectomeSet(InMemoryDataset):
    def __init__(self, root, fname, PATH = None, num_nodes = None, label_name = None, index_name = "Subject", 
                 factorize = False, transform = None, pre_transform = None, pre_filter = None):
        self.fname = fname
        self.PATH = PATH
        self.num_nodes = num_nodes
        self.label_name = label_name
        self.index_name = index_name
        self.factorize = factorize
        super(ConnectomeSet, self).__init__(root, transform, pre_transform, pre_filter)
        self.data, self.slices = torch.load(self.processed_paths[0])

    @property
    def raw_file_names(self):
        return [self.fname]

    @property
    def processed_file_names(self):
        return [self.fname]

    def download(self):
        if self.PATH is None:
            return
        assert len(self.PATH) == 2
        assert self.label_name is not None
        DATA_PATH = self.PATH[0]
        LABEL_PATH = self.PATH[1]
        connectome_process(DATA_PATH, LABEL_PATH, self.raw_dir,
                           self.label_name, self.index_name, self.factorize, output_name=self.fname, num_nodes = self.num_nodes)
        
    def process(self):
        with open(os.path.join(self.raw_dir, self.fname), "rb") as f:
            self.data, self.slices = self.collate(pickle.load(f))

        if self.pre_filter is not None:
            data_list = [self.get(idx) for idx in range(len(self))]
            data_list = [data for data in data_list if self.pre_filter(data)]
            self.data, self.slices = self.collate(data_list)

        if self.pre_transform is not None:
            data_list = [self.get(idx) for idx in range(len(self))]
            data_list = [self.pre_transform(data) for data in data_list]
            self.data, self.slices = self.collate(data_list)

        torch.save((self.data, self.slices), self.processed_paths[0])

    def __repr__(self):
        return '{}()'.format(self.__class__.__name__)
